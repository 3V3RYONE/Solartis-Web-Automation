from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
import openpyxl

def Enrollee_Information(list_data):
    member_number = browser.find_element(by=By.NAME, value="EAANumber")
    member_number.send_keys(list_data[0])

    first_name = browser.find_element(by=By.NAME, value="firstName")
    first_name.send_keys(list_data[1].strip())
   
    if list_data[2] is not None:
        middle_name = browser.find_element(by=By.NAME, value="middleName")
        middle_name.send_keys(list_data[2].strip())

    last_name = browser.find_element(by=By.NAME, value="lastName")
    last_name.send_keys(list_data[3].strip())
    
    DOB = browser.find_element(by=By.NAME, value="insuredDateOfBirth")
    DOB.send_keys(list_data[4].strip())

    Address1 = browser.find_element(by=By.NAME,value="address1")
    Address1.send_keys(list_data[5].strip())

    if list_data[6] is not None:
        Address2 = browser.find_element(by=By.NAME, value="address2")
        Address2.send_keys(list_data[6].strip())

    City = browser.find_element(by=By.NAME, value="city")
    City.send_keys(list_data[7].strip())

    State = Select(browser.find_element(by=By.NAME, value="state"))
    State.select_by_visible_text(list_data[8])

    Zip = browser.find_element(by=By.NAME,value="autocompleteNomeIdInput")
    Zip.send_keys(str(list_data[9]))

    email = browser.find_element(by=By.NAME,value="email")
    email.send_keys(list_data[10].strip())

    phone = browser.find_element(by=By.NAME,value="phone")
    phone.send_keys(list_data[11])

    #enrol_date = browser.find_element_by_name("EnrollmentDateInputDate")
    #work on this to automate this.(list_data[12])

    plan_name = "planNameItem0"
    if 'Spouse' in list_data[13]:
        if 'Children' in list_data[13]:
            plan_name = "planNameItem3"
        else:
            plan_name = "planNameItem1"
    if 'Children' in list_data[13]:
        if 'Spouse' not in list_data[13]:
            plan_name = "planNameItem2"

    plan = browser.find_element(by=By.NAME,value="planNameInput").click()
    browser.find_element(by=By.ID,value=plan_name).click()

    if  list_data[14]==50000:
        sum_value = "coverageLimitItem0"
    if list_data[14]==75000:
        sum_value="coverageLimitItem1"
    elif list_data[14]==100000:
        sum_value="coverageLimitItem2"
    elif list_data[14]==150000:
        sum_value="coverageLimitItem3"
    elif list_data[14]==200000:
        sum_value="coverageLimitItem4"
    elif list_data[14]==250000:
        sum_value="coverageLimitItem5"

    Principal_sum = browser.find_element(by=By.NAME,value="coverageLimitInput").click()
    browser.find_element(by=By.ID, value=sum_value).click()

    Submit_button1 = WebDriverWait(browser,180).until(EC.element_to_be_clickable((By.ID,"submitCertReq2"))).click() #find_element_by_name("submitCertReq2")
    
def Beneficiary(list_ben,count_ben):
   
    for i in range(0,count_ben):
       
        add_ben = WebDriverWait(browser,180000).until(EC.element_to_be_clickable((By.ID,"j_idt66"))).click()
       
        time.sleep(1)
        #first_ben_name = browser.find_element_by_id("beneficiaryFirstName")
        first_ben_name = WebDriverWait(browser,180000).until(EC.presence_of_element_located((By.ID,"beneficiaryFirstName")))
        first_ben_name.send_keys(list_ben[i][0].strip())
       
        time.sleep(1)
        #middle_ben_name = browser.find_element_by_id("beneficiaryMiddleName")
        middle_ben_name = WebDriverWait(browser,180000).until(EC.presence_of_element_located((By.ID,"beneficiaryMiddleName")))

        if list_ben[i][1] is not None:
            middle_ben_name.send_keys(list_ben[i][1].strip())
        else:
            middle_ben_name.send_keys(Keys.TAB)
       
        time.sleep(1)
        #last_ben_name = browser.find_element_by_id("beneficiaryLastName")
        last_ben_name = WebDriverWait(browser,180000).until(EC.presence_of_element_located((By.ID,"beneficiaryLastName")))
        last_ben_name.send_keys(list_ben[i][2].strip())
       
        time.sleep(1)
        relationship = browser.find_element(by=By.ID,value="relationship")
        #relationship = WebDriverWait(browser,30).until(EC.presence_of_element_located((By.ID,"relationship")))
        relationship.send_keys(list_ben[i][3].strip())
        
        percentage = browser.find_element(by=By.ID,value="Beneficiarypercentage")
        percentage.send_keys(list_ben[i][4])
        percentage.clear()
        percentage.send_keys(list_ben[i][4])
       
        save = browser.find_element(by=By.ID, value="j_idt150").click()
        time.sleep(3)
   
    continue_ben = WebDriverWait(browser,180).until(EC.element_to_be_clickable((By.ID,"j_idt68"))).click()

def Dependents(list_dependents,count_dep):
    
    for j in range(0,count_dep):
        
        add_dep = WebDriverWait(browser,180).until(EC.element_to_be_clickable((By.ID,"j_idt62"))).click()

        time.sleep(1)
        first_dep_name = browser.find_element(by=By.ID,value="dependentFirstName")
        first_dep_name.send_keys(list_dependents[j][0].strip())

        if list_dependents[j][1] is not None:
            middle_dep_name = browser.find_element(by=By.ID,value="middleInitial")
            middle_dep_name.send_keys(list_dependents[j][1].strip())

        last_dep_name = browser.find_element(by=By.ID,value="dependentLastName")
        last_dep_name.send_keys(list_dependents[j][2].strip())
        
        DOB_dep = browser.find_element(by=By.ID,value="dependentDateOfBirth")
        DOB_dep.send_keys(list_dependents[j][3].strip())
        
        if list_dependents[j][4].lower()=='spouse':
            relationship_dep = browser.find_element(by=By.ID,value="relationship:0").click()
        elif list_dependents[j][4].lower()=='child':
            relationship_dep = browser.find_element(by=By.ID,value="relationship:1").click()
        if list_dependents[j][5].lower()=='male':
            gender_dep = browser.find_element(by=By.ID,value="gender:0").click()
        elif list_dependents[j][5].lower()=='female':
            gender_dep = browser.find_element(by=By.ID,value="gender:1").click()
        
        save = browser.find_element(by=By.ID,value="j_idt174").click()
    
    continue_dep = WebDriverWait(browser,180).until(EC.element_to_be_clickable((By.ID,"j_idt64"))).click()
    
def Payment(list_payment,list_data):
    
    
    #time.sleep(7)
    temp = WebDriverWait(browser,180).until(EC.element_to_be_clickable((By.ID,"cardNumber"))) #just to load the page 
    card_type = Select(browser.find_element(by=By.ID,value="cardType"))
    #card_type = Select(WebDriverWait(browser,60).until(EC.presence_of_element_located((By.ID,"cardType")))
    card_type.select_by_visible_text(list_payment[1].strip())
    
    card_number = browser.find_element(by=By.ID,value="cardNumber")
    card_number.send_keys(list_payment[2])
    
    list_to_store_expiry_date = list_payment[3].strip().split('/')
    
    expiry_month = Select(browser.find_element(by=By.ID,value="ExpiryMonth"))
    expiry_month.select_by_visible_text(list_to_store_expiry_date[0])
    
    expiry_year = Select(browser.find_element(by=By.ID,value="ExpiryYear"))
    expiry_year.select_by_visible_text(list_to_store_expiry_date[1])
    
    cvv = browser.find_element(by=By.ID,value="cvvNumber")
    cvv.send_keys(list_payment[4])
    
    confirm_mail = browser.find_element(by=By.ID,value="confirmEmailAddress")
    confirm_mail.send_keys(list_data[10].strip())
    
    continue_payment = WebDriverWait(browser,180).until(EC.element_to_be_clickable((By.ID,"j_idt109"))).click()

def Purchase():
    #time.sleep(5)
    if str(sheet.cell(row=r,column=32).value).lower()=='yes':
        agree = WebDriverWait(browser,180).until(EC.element_to_be_clickable((By.ID,"agree"))).click()
        purchase = browser.find_element(by=By.ID,value="pay")
        purchase.click()
        
def Download():
    #time.sleep(7)
    try:
        recipt = WebDriverWait(browser,60).until(EC.element_to_be_clickable((By.ID,"ReceiptPDF")))
        recipt.click()
        COI = browser.find_element(by=By.ID,value="COIPDF")
        COI.click() 
    except:
        return
    
workbook = openpyxl.load_workbook("TestData.xlsx")
sheet = workbook.active
row = sheet.max_row
col = sheet.max_column

options = webdriver.ChromeOptions()
options.add_experimental_option('prefs', {
"download.default_directory": "/home/beleswar/Downloads/Output_selenium", #Change default directory for downloads
"download.prompt_for_download": False, #To auto download the file
"download.directory_upgrade": True,
"plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
})    

r=3
while r<=row:
    browser = webdriver.Chrome(options=options)
    browser.get("https://enrollmentdemo.solartis.net/Quote.xhtml")
    browser.maximize_window()
    time.sleep(2)
    continue_button = browser.find_element(by=By.XPATH,value="/html/body/div/div[2]/form/div/table/tbody/tr/td/table/tbody/tr/td/a/img")
    continue_button.click()
    
    #Enrollee data
    list_data = []
    for c in range(1,16):
        list_data.append(sheet.cell(row=r,column=c).value)
    
    #Beneficiary data
    j=r
    count_ben=0
    list_ben=[]
    while (str(sheet.cell(row=j,column=16).value)!='None'):
        list_temp=[]
        for c in range(16,21):
            temp = sheet.cell(row=j,column=c).value
            list_temp.append(temp)
        list_ben.append(list_temp)    
        count_ben+=1
        j+=1
    
    #Dependents data
    j=r
    count_dep=0
    list_dependents=[]
    while (str(sheet.cell(row=j,column=21).value)!='None'):
        list_temp=[]
        for c in range(21,27):
            temp = sheet.cell(row=j,column=c).value
            list_temp.append(temp)
        list_dependents.append(list_temp)    
        count_dep+=1
        j+=1
    
    #Payment info
    list_payment = []
    for c in range(27,32):
        list_payment.append(sheet.cell(row=r,column=c).value)
        
    #Call functions    
    if str(list_data[13]).lower()=='member only plan':
        Enrollee_Information(list_data)
        Beneficiary(list_ben,count_ben)
        #no dependents
        Payment(list_payment,list_data)
        Purchase()
        # Download()
    else:
        Enrollee_Information(list_data)
        Beneficiary(list_ben,count_ben)
        Dependents(list_dependents,count_dep)
        Payment(list_payment,list_data)
        Purchase()
        # Download()
    
    if count_ben>=count_dep:
        r+=count_ben+1
    else:
        r+=count_dep+1
